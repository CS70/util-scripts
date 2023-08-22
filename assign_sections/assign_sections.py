import random
from enum import Enum

from matcher import Mentor, Preference, Slot, get_matches
from openpyxl import load_workbook
from rich.console import Console, Theme
from rich.table import Table, box

console = Console(theme=Theme({"repr.number": ""}))

SECTION_PREFERENCES_FILE = "preferences.xlsx"
OH_PREFERENCES_FILE = "preferences.xlsx"
SECTION_WORKSHEET_NAME = "Section Matching"
OH_WORKSHEET_NAME = "OH Matching"
SECTION_COUNT_WORKSHEET_NAME = "Section Counts"
OH_COUNT_WORKSHEET_NAME = "OH Counts"

"""Color map for the preferences spreadsheet, in ARGB format."""
COLOR_MAP = {
    # red
    "FFFF0000": 0,
    # orange
    "FFFF9900": 1,
    # yellow
    "FFFFFF00": 3,
    # green
    "FF00FF00": 5,
}
REVERSE_COLOR_MAP = {
    0: "white on #FF0000",
    1: "black on #FF9900",
    3: "black on #FFFF00",
    5: "black on #00FF00",
}


class PreferencesHeader(Enum):
    """Header values for the preferences spreadsheet."""

    LOCATION = "Location"
    DAY = "Day"
    TIME = "Time"
    MIN_COUNT = "Min Count"
    MAX_COUNT = "Max Count"

    # not specified in the spreadsheet; used as a marker for the start of the TA names
    NAMES = "_NAME_START"


class SectionCountHeader(Enum):
    """Header values for the section count spreadsheet."""

    NAME = "Name"
    MIN_COUNT = "Min Count"
    MAX_COUNT = "Max Count"


def discussion_to_str(discussion):
    """
    Convert a discussion info dict into a human-readable string.
    """
    return f"{discussion['day']} {discussion['time']} @ {discussion['location']}"


def load_excel(filename: str, sheet_name: str):
    """
    Load the excel file and parse colors into values.

    Returns two dictionaries:
    - discussions
        Map from the row number to the discussion description
    - preferences
        Map from TA names to their discussion preferences;
        the discussion preferences for a TA are given in a dict,
        from the row number to the preference for that discussion.

    Getting the link from the discussion description to the TA preference
    should be done through both dictionaries; the shared ID is through the row number,
    to ensure unique IDs for each possible discussion slot.
    """
    workbook = load_workbook(filename=filename)
    worksheet = workbook[sheet_name]

    # map from type to column number
    column_map = {}

    # total number of columns to look at
    num_cols = 0
    # whether we're currently parsing the metadata;
    # once we encounter a header value we do not recognize, we assume it is a TA's name.
    parsing_metadata = True
    for col in worksheet.iter_cols():
        cur_header = col[0].value
        # current column index
        cur_col = col[0].column
        if not cur_header:
            break

        # keep track of the number of filled columns
        num_cols = cur_col

        if not parsing_metadata:
            # only overwrite column map if we are parsing metadata
            pass
        elif cur_header == PreferencesHeader.LOCATION.value:
            column_map[PreferencesHeader.LOCATION] = cur_col
        elif cur_header == PreferencesHeader.DAY.value:
            column_map[PreferencesHeader.DAY] = cur_col
        elif cur_header == PreferencesHeader.TIME.value:
            column_map[PreferencesHeader.TIME] = cur_col
        elif cur_header == PreferencesHeader.MIN_COUNT.value:
            column_map[PreferencesHeader.MIN_COUNT] = cur_col
        elif cur_header == PreferencesHeader.MAX_COUNT.value:
            column_map[PreferencesHeader.MAX_COUNT] = cur_col
        else:
            # unrecognized header value; stop parsing metadata
            parsing_metadata = False
            column_map[PreferencesHeader.NAMES] = cur_col

            # validate metadata values
            missing = [
                enum.value
                for enum in (
                    PreferencesHeader.LOCATION,
                    PreferencesHeader.DAY,
                    PreferencesHeader.TIME,
                )
                if enum not in column_map
            ]
            if missing:
                raise ValueError(
                    f"Invalid preference sheet headers; missing {missing}."
                    " Ensure that all metadata fields precede course staff names."
                )

    assert num_cols > 0

    # map from row number to discussion description
    discussions = {}
    num_rows = 0
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value:
            num_rows = row[0].row
        else:
            # stop when we first see an empty cell
            break

        # look up column and store metadata;
        # subtracting 1 from the values in the column map, since Excel is 1-indexed.
        discussions[row[0].row] = {
            "location": row[column_map[PreferencesHeader.LOCATION] - 1].value,
            "day": row[column_map[PreferencesHeader.DAY] - 1].value,
            "time": row[column_map[PreferencesHeader.TIME] - 1].value,
            "min": (
                row[column_map[PreferencesHeader.MIN_COUNT] - 1].value
                if PreferencesHeader.MIN_COUNT in column_map
                # default to 0 if not specified
                else 0
            ),
            "max": (
                row[column_map[PreferencesHeader.MAX_COUNT] - 1].value
                if PreferencesHeader.MAX_COUNT in column_map
                # default to 1 if not specified
                else 1
            ),
        }

    preferences = {}
    for col in worksheet.iter_cols(
        min_col=column_map[PreferencesHeader.NAMES], max_col=num_cols
    ):
        ta_preferences = {}
        for cell in col[1:num_rows]:
            pref_col = cell.fill.bgColor.rgb
            assert pref_col in COLOR_MAP, f"Invalid preference color (ARGB): {pref_col}"

            pref = COLOR_MAP[pref_col]
            ta_preferences[cell.row] = pref

        # col[0] is the header containing the TA name
        preferences[col[0].value] = ta_preferences

    return discussions, preferences


def load_num_sections(filename: str, sheet_name: str):
    """
    Load the number of sections that we should match per TA.
    """
    workbook = load_workbook(filename=filename)
    worksheet = workbook[sheet_name]

    # scan columns for names
    column_map = {}
    for col in worksheet.iter_cols():
        cur_header = col[0].value
        cur_col = col[0].column

        if not col[0].value:
            break

        if cur_header == SectionCountHeader.NAME.value:
            column_map[SectionCountHeader.NAME] = cur_col
        elif cur_header == SectionCountHeader.MIN_COUNT.value:
            column_map[SectionCountHeader.MIN_COUNT] = cur_col
        elif cur_header == SectionCountHeader.MAX_COUNT.value:
            column_map[SectionCountHeader.MAX_COUNT] = cur_col
        else:
            # unrecognized; error
            raise ValueError(f"Unrecognized header: {cur_header}")

    # validate column map
    missing = [
        enum.value
        for enum in (
            SectionCountHeader.NAME,
            SectionCountHeader.MIN_COUNT,
            SectionCountHeader.MAX_COUNT,
        )
        if enum not in column_map
    ]
    if missing:
        raise ValueError(f"Invalid section count sheet headers; missing {missing}")

    num_sections = {}
    for row in worksheet.iter_rows(min_row=2):
        if not row[0].value:
            # stop when we first see an empty cell
            break

        # get cell values; subtracting 1 since Excel is 1-indexed
        name = row[column_map[SectionCountHeader.NAME] - 1].value
        min_count = int(row[column_map[SectionCountHeader.MIN_COUNT] - 1].value)
        max_count = int(row[column_map[SectionCountHeader.MAX_COUNT] - 1].value)

        # keep track of these values
        num_sections[name] = {
            "min": min_count,
            "max": max_count,
        }

    return num_sections


def run_matcher(
    preferences_file: str, preference_worksheet_name: str, count_worksheet_name: str
):
    """
    Run the matcher on a given preferences file, worksheet name, and section count worksheet name.

    `preferences_file`:
        An Excel file that will be used to fetch all preferences and section count info.
    `preference_worksheet_name`:
        Used to fetch the preferences for each TA.
    `count_worksheet_name`:
        Used to fetch the section count for each TA.
    """
    slot_id_to_info, preference_map = load_excel(
        preferences_file, preference_worksheet_name
    )
    num_sections_map = load_num_sections(preferences_file, count_worksheet_name)

    # format input values
    mentors = [
        Mentor(
            id=name,
            name=name,
            min_slots=slots["min"],
            max_slots=slots["max"],
        )
        for name, slots in num_sections_map.items()
    ]
    slots = [
        Slot(
            id=row,
            time=f"{slot_info['day']} {slot_info['time']}",
            location=slot_info["location"],
            min_mentors=slot_info["min"],
            max_mentors=slot_info["max"],
        )
        for row, slot_info in slot_id_to_info.items()
    ]
    preferences = [
        Preference(mentor_id=mentor, slot_id=row, value=value)
        for mentor, pref in preference_map.items()
        for row, value in pref.items()
    ]
    random.shuffle(mentors)

    result = get_matches(mentors, slots, preferences)
    cost = result["cost"]
    assignments = result["assignments"]
    unmatched = result["unmatched"]
    print("cost", cost)
    print("unmatched", unmatched)

    assigned_by_slot = {}

    # print table by TA
    table_rows = []
    for name, assigned in sorted(assignments.items(), key=lambda t: t[0]):
        formatted_discussions = []
        for row in sorted(assigned):
            disc_str = discussion_to_str(slot_id_to_info[row])
            pref = preference_map[name][row]
            pref_color = REVERSE_COLOR_MAP[pref]
            formatted_discussions.append(f"[{pref_color}]{disc_str}[/{pref_color}]")

            # restructure the dict
            if row not in assigned_by_slot:
                assigned_by_slot[row] = set()
            assigned_by_slot[row].add(name)

        table_rows.append([name, *formatted_discussions])

    # name column, followed by assigned slots
    table_by_ta = Table("Name", "Assigned", box=box.SIMPLE)
    num_columns = max(len(row) for row in table_rows)
    for _ in range(num_columns - 2):
        table_by_ta.add_column()

    for table_row in table_rows:
        if len(table_row) != num_columns:
            table_row = [*table_row, *([""] * (num_columns - len(table_row)))]
        table_by_ta.add_row(*table_row)
    console.print(table_by_ta)

    # print table by slot
    table_rows = []
    for row, discussion_info in slot_id_to_info.items():
        tas = sorted(assigned_by_slot[row])

        colored_tas = []
        for name in tas:
            pref = preference_map[name][row]
            pref_color = REVERSE_COLOR_MAP[pref]
            colored_tas.append(f"[{pref_color}]{name}[/{pref_color}]")
        table_rows.append(
            [
                discussion_info["location"],
                discussion_info["day"],
                discussion_info["time"],
                *colored_tas,
            ]
        )

    # location, day, time columns followed by assigned TAs
    table_by_slot = Table("Location", "Day", "Time", "Assigned", box=box.SIMPLE)
    num_columns = max(len(row) for row in table_rows)
    for _ in range(num_columns - 4):
        table_by_ta.add_column()

    for table_row in table_rows:
        if len(table_row) != num_columns:
            table_row = [*table_row, *([""] * (num_columns - len(table_row)))]
        table_by_slot.add_row(*table_row)
    console.print(table_by_slot)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--section", action="store_true", help="Match sections")
    group.add_argument("--oh", action="store_true", help="Match OH")
    parser.add_argument("--seed", "-s", help="Random seed")
    args = parser.parse_args()

    if args.seed:
        random.seed(args.seed)

    if args.section:
        run_matcher(
            SECTION_PREFERENCES_FILE,
            SECTION_WORKSHEET_NAME,
            SECTION_COUNT_WORKSHEET_NAME,
        )
    elif args.oh:
        run_matcher(OH_PREFERENCES_FILE, OH_WORKSHEET_NAME, OH_COUNT_WORKSHEET_NAME)
