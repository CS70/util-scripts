"""
Converts a spreadsheet of preferences and settings into CSV files,
for standard use with the matcher.
"""

import csv
import json
from typing import Any, TypedDict, cast

from openpyxl import load_workbook

# DEFAULTS
SECTION_WORKSHEET_NAME = "Section Matching"
OH_WORKSHEET_NAME = "OH Matching"
SECTION_COUNT_WORKSHEET_NAME = "Section Counts"
OH_COUNT_WORKSHEET_NAME = "OH Counts"

SECTION_OUT_FILENAME = "section_preferences.csv"
OH_OUT_FILENAME = "oh_preferences.csv"
SECTION_CONFIG_OUT_FILENAME = "section_config.json"
OH_CONFIG_OUT_FILENAME = "oh_config.json"


COLOR_MAP = {
    # red
    "FFFF0000": 0,
    # orange
    "FFFF9900": 1,
    # yellow
    "FFFFFF00": 3,
    # green
    "FF00FF00": 5,
}
"""Color map for the preferences spreadsheet, in ARGB format."""

# ===== TYPES =====


class PreferencesHeader:
    """Header values for the preferences spreadsheet."""

    LOCATION = "Location"
    DAY = "Day"
    START_TIME = "Start Time"
    END_TIME = "End Time"
    MIN_COUNT = "Min Count"
    MAX_COUNT = "Max Count"

    # not in spreadsheet, but will be added to the CSV for configuration purposes
    ID = "ID"

    # not specified in the spreadsheet; used as a marker for the start of the TA names
    NAMES = "_NAME_START"


class SlotCountHeader:
    """Header values for the section count spreadsheet."""

    NAME = "Name"
    MIN_COUNT = "Min Count"
    MAX_COUNT = "Max Count"


DiscussionInfo = TypedDict(
    "DiscussionInfo",
    {
        "ID": str,
        "Location": str,
        "Day": str,
        "Start Time": str,
        "End Time": str,
        "Min Count": int,
        "Max Count": int,
    },
)
RawDiscussionMap = dict[int, DiscussionInfo]

PreferencesMap = dict[str, dict[int, int]]

SectionCountInfo = TypedDict("SectionCountInfo", {"Min Count": int, "Max Count": int})
SectionCountMap = dict[str, SectionCountInfo]


def slot_name(info: DiscussionInfo):
    location = info[PreferencesHeader.LOCATION]
    day = info[PreferencesHeader.DAY]
    start_time = info[PreferencesHeader.START_TIME]
    end_time = info[PreferencesHeader.END_TIME]

    return f"{location}|{day}|{start_time}|{end_time}"


def load_excel(
    filename: str, sheet_name: str
) -> tuple[RawDiscussionMap, PreferencesMap]:
    """
    Load the excel file and parse colors into values.

    Returns two dictionaries:
    - discussions
        Map from the row number to the discussion description
    - preferences
        Map from TA names to their discussion preferences;
        the discussion preferences for a TA are given in a dict,
        from the row number to the preference for that discussion.

    Getting the link from the discussion description to the TA preference
    should be done through both dictionaries; the shared ID is through the row number,
    to ensure unique IDs for each possible discussion slot.
    """
    workbook = load_workbook(filename=filename)
    worksheet = workbook[sheet_name]

    # map from type to column number
    column_map = {}

    # total number of columns to look at
    num_cols = 0
    # whether we're currently parsing the metadata;
    # once we encounter a header value we do not recognize, we assume it is a TA's name.
    parsing_metadata = True
    for col in worksheet.iter_cols():
        cur_header = col[0].value
        # current column index
        cur_col = col[0].column
        if not cur_header:
            break

        # keep track of the number of filled columns
        num_cols = cur_col

        if not parsing_metadata:
            # only overwrite column map if we are parsing metadata
            pass
        elif cur_header == PreferencesHeader.LOCATION:
            column_map[PreferencesHeader.LOCATION] = cur_col
        elif cur_header == PreferencesHeader.DAY:
            column_map[PreferencesHeader.DAY] = cur_col
        elif cur_header == PreferencesHeader.START_TIME:
            column_map[PreferencesHeader.START_TIME] = cur_col
        elif cur_header == PreferencesHeader.END_TIME:
            column_map[PreferencesHeader.END_TIME] = cur_col
        elif cur_header == PreferencesHeader.MIN_COUNT:
            column_map[PreferencesHeader.MIN_COUNT] = cur_col
        elif cur_header == PreferencesHeader.MAX_COUNT:
            column_map[PreferencesHeader.MAX_COUNT] = cur_col
        else:
            # unrecognized header value; stop parsing metadata
            parsing_metadata = False
            column_map[PreferencesHeader.NAMES] = cur_col

            # validate metadata values
            missing = [
                enum
                for enum in (
                    PreferencesHeader.LOCATION,
                    PreferencesHeader.DAY,
                    PreferencesHeader.START_TIME,
                    PreferencesHeader.END_TIME,
                )
                if enum not in column_map
            ]
            if missing:
                raise ValueError(
                    f"Invalid preference sheet headers; missing {missing}."
                    " Ensure that all metadata fields precede course staff names."
                )

    assert num_cols > 0

    # map from row number to discussion description
    discussions: RawDiscussionMap = {}
    num_rows = 0
    for row in worksheet.iter_rows(min_row=2):
        # slot ID should be 0-indexed
        slot_id = row[0].row - 2

        if row[0].value:
            num_rows = row[0].row
        else:
            # stop when we first see an empty cell
            break

        # look up column and store metadata;
        # subtracting 1 from the values in the column map, since Excel is 1-indexed.
        discussions[slot_id] = cast(
            DiscussionInfo,
            {
                PreferencesHeader.LOCATION: row[
                    column_map[PreferencesHeader.LOCATION] - 1
                ].value,
                PreferencesHeader.DAY: row[column_map[PreferencesHeader.DAY] - 1].value,
                PreferencesHeader.START_TIME: row[
                    column_map[PreferencesHeader.START_TIME] - 1
                ].value,
                PreferencesHeader.END_TIME: row[
                    column_map[PreferencesHeader.END_TIME] - 1
                ].value,
                PreferencesHeader.MIN_COUNT: (
                    int(row[column_map[PreferencesHeader.MIN_COUNT] - 1].value)
                    if PreferencesHeader.MIN_COUNT in column_map
                    # default to 0 if not specified
                    else 0
                ),
                PreferencesHeader.MAX_COUNT: (
                    int(row[column_map[PreferencesHeader.MAX_COUNT] - 1].value)
                    if PreferencesHeader.MAX_COUNT in column_map
                    # default to 1 if not specified
                    else 1
                ),
            },
        )

    preferences = {}
    for col in worksheet.iter_cols(
        min_col=column_map[PreferencesHeader.NAMES], max_col=num_cols
    ):
        user_preferences = {}
        for cell in col[1:num_rows]:
            # slot id is 0-indexed by row
            slot_id = cell.row - 2

            pref_col = cell.fill.bgColor.rgb
            assert pref_col in COLOR_MAP, f"Invalid preference color (ARGB): {pref_col}"

            pref = COLOR_MAP[pref_col]
            user_preferences[slot_id] = pref

        # col[0] is the header containing the TA name
        preferences[col[0].value] = user_preferences

    return discussions, preferences


def load_num_sections(filename: str, sheet_name: str) -> SectionCountMap:
    """
    Load the number of sections that we should match per user.
    """
    workbook = load_workbook(filename=filename)
    worksheet = workbook[sheet_name]

    # scan columns for names
    column_map = {}
    for col in worksheet.iter_cols():
        cur_header = col[0].value
        cur_col = col[0].column

        if not col[0].value:
            break

        if cur_header == SlotCountHeader.NAME:
            column_map[SlotCountHeader.NAME] = cur_col
        elif cur_header == SlotCountHeader.MIN_COUNT:
            column_map[SlotCountHeader.MIN_COUNT] = cur_col
        elif cur_header == SlotCountHeader.MAX_COUNT:
            column_map[SlotCountHeader.MAX_COUNT] = cur_col
        else:
            # unrecognized; error
            raise ValueError(f"Unrecognized header: {cur_header}")

    # validate column map
    missing = [
        enum
        for enum in (
            SlotCountHeader.NAME,
            SlotCountHeader.MIN_COUNT,
            SlotCountHeader.MAX_COUNT,
        )
        if enum not in column_map
    ]
    if missing:
        raise ValueError(f"Invalid section count sheet headers; missing {missing}")

    num_sections = {}
    for row in worksheet.iter_rows(min_row=2):
        if not row[0].value:
            # stop when we first see an empty cell
            break

        # get cell values; subtracting 1 since Excel is 1-indexed
        name = row[column_map[SlotCountHeader.NAME] - 1].value
        min_count = int(row[column_map[SlotCountHeader.MIN_COUNT] - 1].value)
        max_count = int(row[column_map[SlotCountHeader.MAX_COUNT] - 1].value)

        # keep track of these values
        num_sections[name] = {
            SlotCountHeader.MIN_COUNT: min_count,
            SlotCountHeader.MAX_COUNT: max_count,
        }

    return num_sections


def convert_sheet(
    spreadsheet_filename,
    preferences_sheet_name,
    count_sheet_name,
    csv_output_filename,
    json_output_filename,
):
    # parse sheet
    info_map, preferences = load_excel(spreadsheet_filename, preferences_sheet_name)
    num_slots_map = load_num_sections(spreadsheet_filename, count_sheet_name)

    # write preference CSV
    with open(csv_output_filename, "w", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                PreferencesHeader.ID,
                PreferencesHeader.LOCATION,
                PreferencesHeader.DAY,
                PreferencesHeader.START_TIME,
                PreferencesHeader.END_TIME,
                *preferences.keys(),
            ],
        )
        writer.writeheader()

        for slot_id, info in info_map.items():
            row: Any = info.copy()
            # add ID to the csv row
            row[PreferencesHeader.ID] = slot_id
            # remove counts from the csv row
            del row[PreferencesHeader.MIN_COUNT]
            del row[PreferencesHeader.MAX_COUNT]

            for name, pref_map in preferences.items():
                row[name] = pref_map[slot_id]

            writer.writerow(row)

    # compile configuration object
    config = {
        "users": {
            name: {
                "min_slots": info[SlotCountHeader.MIN_COUNT],
                "max_slots": info[SlotCountHeader.MAX_COUNT],
            }
            for name, info in num_slots_map.items()
        },
        "slots": {
            slot_id: {
                "min_users": info[PreferencesHeader.MIN_COUNT],
                "max_users": info[PreferencesHeader.MAX_COUNT],
            }
            for slot_id, info in info_map.items()
        },
    }

    # write configuration JSON
    with open(json_output_filename, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)


def main(
    spreadsheet_filename,
    section_sheet_name=SECTION_WORKSHEET_NAME,
    oh_sheet_name=OH_WORKSHEET_NAME,
    section_count_sheet_name=SECTION_COUNT_WORKSHEET_NAME,
    oh_count_sheet_name=OH_COUNT_WORKSHEET_NAME,
    section_out=SECTION_OUT_FILENAME,
    oh_out=OH_OUT_FILENAME,
    section_config_out=SECTION_CONFIG_OUT_FILENAME,
    oh_config_out=OH_CONFIG_OUT_FILENAME,
):
    """
    Main function, converting both the section and OH preferences into CSVs.
    """

    convert_sheet(
        spreadsheet_filename,
        section_sheet_name,
        section_count_sheet_name,
        section_out,
        section_config_out,
    )
    convert_sheet(
        spreadsheet_filename, oh_sheet_name, oh_count_sheet_name, oh_out, oh_config_out
    )


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("spreadsheet_file", help="Spreadsheet to convert")

    name_group = parser.add_argument_group("Sheet name options")
    name_group.add_argument(
        "--section-sheet-name",
        default=SECTION_WORKSHEET_NAME,
        help="Name of the sheet for section preferences",
    )
    name_group.add_argument(
        "--oh-sheet-name",
        default=OH_WORKSHEET_NAME,
        help="Name of the sheet for OH preferences",
    )

    out_group = parser.add_argument_group("Output options")
    out_group.add_argument(
        "--section-out",
        default=SECTION_OUT_FILENAME,
        help="Output file for discussion preferences",
    )
    out_group.add_argument(
        "--oh-out", default=OH_OUT_FILENAME, help="Output file for OH preferences"
    )

    args = parser.parse_args()
    main(args.spreadsheet_file, section_out=args.section_out, oh_out=args.oh_out)
